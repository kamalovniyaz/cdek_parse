import os
import re
import time

import openpyxl
from playwright.sync_api import Playwright, sync_playwright


def watch_excel():
    """Открывает и просматривает excel файл, получая нужную информацию и создавая массив данных для дальнейшей обработки"""

    excel_file_path = os.path.join('excel_files', 'cdek_prices.xlsx')
    file_to_read = openpyxl.load_workbook(excel_file_path, data_only=True)
    sheet = file_to_read["экспресс"]
    weights = []
    cityes = []
    for row in range(4, 25):
        ship = []
        for col in range(2, 4):
            value = sheet.cell(row, col).value
            ship.append(value)
        ship.append({'row': row})
        cityes.append([{'a_point': ship[0], 'b_point': ship[1], 'row': row}])

    for ship_row in range(3, 4):
        for ship_col in range(4, 11):
            weight = sheet.cell(ship_row, ship_col).value
            weights.append([{'weight': weight, 'col': ship_col}])

    return cityes, weights


def run_checking_prices(cityes, weights):
    """Запуск playwright в синхронном режиме"""

    with sync_playwright() as playwright:
        checking_prices(playwright, cityes, weights)
        return {'status': 'success'}


def checking_prices(playwright: Playwright, cityes, weights) -> None:
    """Запускает браузер и выполняет действия для парсинга цен основываясь на данные excel файла"""

    try:
        browser = playwright.chromium.launch(headless=False)
        context = browser.new_context()
        page = context.new_page()
        excel_file_path = os.path.join('excel_files', 'cdek_prices.xlsx')
        file_to_write = openpyxl.load_workbook(excel_file_path)
        sheet = file_to_write["экспресс"]

        for i in range(len(cityes)):
            datesWrite = False
            for j in range(len(weights)):
                try:
                    page.goto('https://www.cdek.ru/ru/cabinet/calculate')
                    time.sleep(5)
                    page.get_by_role("banner").get_by_role("button", name="Отправить или получить").click()
                    page.get_by_role("link", name="Отправить посылку Рассчитайте и оформите накладную онлайн").click()

                    a_point = cityes[i][0].get('a_point')
                    b_point = cityes[i][0].get('b_point')
                    weight = weights[j][0].get('weight')

                    page.get_by_text("Город отправки").click()
                    page.get_by_label("Город отправки").fill(a_point)
                    time.sleep(2)
                    page.wait_for_selector('.cdek-dropdown-item').click()
                    page.get_by_text("Город назначения").click()
                    page.get_by_label("Город назначения").fill(b_point)
                    time.sleep(2)
                    page.wait_for_selector('.cdek-dropdown-item').click()
                    page.get_by_text("Размер посылки").click()
                    elements = page.query_selector_all(".choice-of-dimensions-item__data")

                    for element in elements:
                        # Получите текст из элементов
                        text = element.text_content().lower()
                        if "короб" in text and f"{weight} кг" in text:
                            element.click()
                            break

                    page.get_by_role("button", name="Рассчитать").click()
                    page.wait_for_selector('.main-tabs__period')
                    faster_ship = page.get_by_text('Экспресс-доставка ко времени')
                    faster_ship.click()
                    time.sleep(2)

                    faster_ship_dates = page.query_selector('.info-order__days-text').text_content()
                    faster_ship_price = page.query_selector('.info-sum__amount').text_content()

                    if '-' in faster_ship_dates:
                        from_days, to_days = faster_ship_dates.split('-')
                        to_days = re.sub(r'\D', '', to_days)
                    else:
                        day = re.sub(r'\D', '', faster_ship_dates)
                        from_days = 0
                        to_days = day

                    sheet.cell(row=cityes[i][0].get('row'), column=weights[j][0].get('col'),
                               value=faster_ship_price)

                    if not datesWrite:
                        sheet.cell(row=cityes[i][0].get('row'), column=11, value=from_days)
                        sheet.cell(row=cityes[i][0].get('row'), column=12, value=to_days)
                    time.sleep(2)

                    file_to_write.save('excel_files/cdek_prices.xlsx')

                except Exception as f:
                    pass
        browser.close()

        file_to_write.close()
    except Exception as ex:
        pass


if __name__ == '__main__':
    cityes, weights = watch_excel()
    run_checking_prices(cityes, weights)
